from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import uuid
import os
import datetime
import json
import io
import oracledb
from sqlalchemy import create_engine, text
import random
import re
import traceback
from modules.login.login import auth_bp, token_required
from modules.admin.admin import admin_bp, admin_required
from dotenv import load_dotenv
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database.dbconnect import create_oracle_connection, sqlite_engine

# Import blueprints
from modules.license.license import license_bp
from modules.mapper.mapper import mapper_bp
from modules.jobs.jobs import jobs_bp
from modules.type_mapping.parameter_mapping import parameter_mapping_bp
from modules.dashboard.dashboard import dashboard_bp

app = Flask(__name__)
# CORS(app, resources={
#     r"/*": {
#         "origins": ["http://localhost:3000"],
#         "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
#         "allow_headers": ["Content-Type", "Authorization"],
#         "supports_credentials": True,
#         "expose_headers": ["Authorization"]
#     }
# })

# allow all origins with  support credentials
CORS(app, resources={
    r"/*": {
        "origins": ["*"],
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "supports_credentials": True,
    }
})


# Load environment variables
load_dotenv()

# Register blueprints
app.register_blueprint(auth_bp, url_prefix='/auth')
app.register_blueprint(admin_bp, url_prefix='/admin')
app.register_blueprint(license_bp, url_prefix='/api')
app.register_blueprint(mapper_bp, url_prefix='/mapper')
app.register_blueprint(jobs_bp, url_prefix='/job')
app.register_blueprint(parameter_mapping_bp, url_prefix='/mapping')
app.register_blueprint(dashboard_bp, url_prefix='/dashboard')

# Create directories if they don't exist
os.makedirs('data/drafts', exist_ok=True)
os.makedirs('data/templates', exist_ok=True)

if __name__ == '__main__':
    app.run(debug=True, port=5000)   
 