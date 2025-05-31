from flask import Blueprint, request, jsonify, send_file
from database.dbconnect import create_oracle_connection
import pandas as pd
import uuid
import os
import datetime
import json
import io
import oracledb
import re
import traceback
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from modules.helper_functions import create_update_mapping,create_update_mapping_detail, validate_logic2, validate_all_mapping_details,  get_mapping_ref  ,get_mapping_details,get_error_messages_list,get_parameter_mapping_datatype,get_parameter_mapping_scd_type,call_activate_deactivate_mapping, call_delete_mapping, call_delete_mapping_details, call_schedule_immediate_job


# Create blueprint
mapper_bp = Blueprint('mapper', __name__)

FORM_FIELDS = ["reference","description","sourceSystem","tableName","tableType","targetSchema","freqCode","bulkProcessRows"]
TABLE_FIELDS = ["primaryKey","pkSeq","fieldName","dataType","fieldDesc","scdType","keyColumn","valColumn","logic","mapCombineCode","execSequence"]


# routes

@mapper_bp.route('/download-template', methods=['GET'])
def download_template():
    try:
        # Check if format parameter is provided, default to xlsx
        export_format = request.args.get('format', 'xlsx').lower()
        if export_format not in ['xlsx', 'csv']:
            export_format = 'xlsx'  # Default to xlsx if format is invalid
            
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Mapping Template"
 
        # Define styles
        header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
 
        # Write Form Fields section
        ws['A1'] = "Form Fields"
        ws.merge_cells('A1:H1')
        ws['A1'].fill = header_fill
        ws['A1'].font = header_font
        ws['A1'].alignment = header_alignment
 
        # Write Form Fields headers
        for col, field in enumerate(FORM_FIELDS, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = field
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
 
        # Add empty row for form data
        for col in range(1, len(FORM_FIELDS) + 1):
            cell = ws.cell(row=3, column=col)
            cell.border = border
 
        # Add space between sections
        ws.append([])
 
        # Write Table Fields section
        current_row = 5
        ws.cell(row=current_row, column=1, value="Table Fields")
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws.cell(row=current_row, column=1).fill = header_fill
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=1).alignment = header_alignment
 
        # Write Table Fields headers
        current_row += 1
        for col, field in enumerate(TABLE_FIELDS, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = field
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
 
        # Add empty rows for table data
        for row in range(current_row + 1, current_row + 11):  # 10 empty rows
            for col in range(1, len(TABLE_FIELDS) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
 
        # Adjust column widths
        for col_idx in range(1, max(len(FORM_FIELDS), len(TABLE_FIELDS)) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
 
            # Check form headers (row 2)
            if col_idx <= len(FORM_FIELDS):
                header_value = ws.cell(row=2, column=col_idx).value
                if header_value:
                    max_length = max(max_length, len(str(header_value)))
 
            # Check table headers (row 6)
            if col_idx <= len(TABLE_FIELDS):
                header_value = ws.cell(row=current_row, column=col_idx).value
                if header_value:
                    max_length = max(max_length, len(str(header_value)))
 
            # Set the column width
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
 
        # Save to BytesIO buffer
        excel_buffer = io.BytesIO()
        
        if export_format == 'csv':
            # Create DataFrame from workbook data to save as CSV
            form_headers = [ws.cell(row=2, column=i).value for i in range(1, len(FORM_FIELDS) + 1)]
            form_data = [ws.cell(row=3, column=i).value or '' for i in range(1, len(FORM_FIELDS) + 1)]
            
            table_headers = [ws.cell(row=current_row, column=i).value for i in range(1, len(TABLE_FIELDS) + 1)]
            table_data = []
            for row in range(current_row + 1, current_row + 11):
                row_data = []
                for col in range(1, len(TABLE_FIELDS) + 1):
                    row_data.append(ws.cell(row=row, column=col).value or '')
                table_data.append(row_data)
            
            # First save form data
            form_df = pd.DataFrame([form_data], columns=form_headers)
            
            # Then save table data
            table_df = pd.DataFrame(table_data, columns=table_headers)
            
            # Combine them with a separator row
            combined_csv = form_df.to_csv(index=False) + "\n\n" + table_df.to_csv(index=False)
            
            # Convert to bytes and prepare response
            excel_buffer = io.BytesIO(combined_csv.encode())
            mimetype = 'text/csv'
            download_name = 'mapper_template.csv'
        else:  # xlsx format
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            download_name = 'mapper_template.xlsx'
 
        return send_file(
            excel_buffer,
            mimetype=mimetype,
            as_attachment=True,
            download_name=download_name
        )
    except Exception as e:
        print(f"Error in download_template: {str(e)}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@mapper_bp.route('/download-current', methods=['POST'])
def download_current():
    try:
        data = request.json
        form_data = data.get('formData', {})
        rows_data = data.get('rows', [])
        
        # Check if format parameter is provided, default to xlsx
        export_format = data.get('format', 'xlsx').lower()
        if export_format not in ['xlsx', 'csv']:
            export_format = 'xlsx'  # Default to xlsx if format is invalid
            
        # Filter out rows without field names
        rows_data = [row for row in rows_data if row.get('fieldName')]
            
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Mapping Template"
 
        # Define styles
        header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
 
        # Write Form Fields section
        ws['A1'] = "Form Fields"
        ws.merge_cells('A1:H1')
        ws['A1'].fill = header_fill
        ws['A1'].font = header_font
        ws['A1'].alignment = header_alignment
 
        # Write Form Fields headers
        for col, field in enumerate(FORM_FIELDS, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = field
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
 
        # Write Form Fields values
        for col, field in enumerate(FORM_FIELDS, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = form_data.get(field, '')
            cell.border = border
 
        # Add space between sections
        ws.append([])
 
        # Write Table Fields section
        current_row = 5
        ws.cell(row=current_row, column=1, value="Table Fields")
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws.cell(row=current_row, column=1).fill = header_fill
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=1).alignment = header_alignment
 
        # Write Table Fields headers
        current_row += 1
        for col, field in enumerate(TABLE_FIELDS, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = field
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
 
        # Add rows with data
        for row_idx, row_data in enumerate(rows_data, current_row + 1):
            # Map the field names to match TABLE_FIELDS
            field_mapping = {
                'primaryKey': 'primaryKey',
                'pkSeq': 'pkSeq',
                'fieldName': 'fieldName',
                'dataType': 'dataType',
                'fieldDesc': 'fieldDesc',
                'scdType': 'scdType',
                'keyColumn': 'keyColumn',
                'valColumn': 'valColumn',
                'logic': 'logic',
                'mapCombineCode': 'mapCombineCode',
                'execSequence': 'execSequence',
            }
            
            for col_idx, field in enumerate(TABLE_FIELDS, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Get the mapped field name
                mapped_field = field_mapping.get(field, field)
                
                # Handle boolean fields like primaryKey separately
                if field == 'primaryKey':
                    cell.value = 'Yes' if row_data.get(mapped_field, False) else 'No'
                else:
                    cell.value = row_data.get(mapped_field, '')
                
                cell.border = border
 
        # Adjust column widths
        for col_idx in range(1, max(len(FORM_FIELDS), len(TABLE_FIELDS)) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
 
            # Check all rows for maximum value length
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    cell_value_str = str(cell_value)
                    # Limit preview length for very long values
                    max_length = max(max_length, min(len(cell_value_str), 50))
 
            # Set the column width
            adjusted_width = max(max_length + 2, 12)  # Minimum width of 12
            ws.column_dimensions[column_letter].width = adjusted_width
 
        # Save to BytesIO buffer
        excel_buffer = io.BytesIO()
        
        if export_format == 'csv':
            # Create DataFrame from workbook data to save as CSV
            form_headers = [ws.cell(row=2, column=i).value for i in range(1, len(FORM_FIELDS) + 1)]
            form_data_values = [ws.cell(row=3, column=i).value or '' for i in range(1, len(FORM_FIELDS) + 1)]
            
            table_headers = [ws.cell(row=current_row, column=i).value for i in range(1, len(TABLE_FIELDS) + 1)]
            table_data = []
            for row in range(current_row + 1, current_row + 1 + len(rows_data)):
                row_data = []
                for col in range(1, len(TABLE_FIELDS) + 1):
                    row_data.append(ws.cell(row=row, column=col).value or '')
                table_data.append(row_data)
            
            # First save form data
            form_df = pd.DataFrame([form_data_values], columns=form_headers)
            
            # Then save table data
            table_df = pd.DataFrame(table_data, columns=table_headers)
            
            # Combine them with a separator row
            combined_csv = form_df.to_csv(index=False) + "\n\n" + table_df.to_csv(index=False)
            
            # Convert to bytes and prepare response
            excel_buffer = io.BytesIO(combined_csv.encode())
            mimetype = 'text/csv'
            download_name = f"{form_data.get('reference', 'mapper')}_data.csv"
        else:  # xlsx format
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            download_name = f"{form_data.get('reference', 'mapper')}_data.xlsx"
 
        return send_file(
            excel_buffer,
            mimetype=mimetype,
            as_attachment=True,
            download_name=download_name
        )
        
    except Exception as e:
        print(f"Error in download_current: {str(e)}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@mapper_bp.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
 
        file = request.files['file']
        # print(file)
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
 
        # Read the Excel file
        wb = load_workbook(io.BytesIO(file.read()))
        ws = wb.active
 
        # Process form fields
        form_data = {}
        form_headers = [cell.value for cell in ws[2] if cell.value]  # Get headers from row 2
        form_values = [cell.value for cell in ws[3] if cell.column <= len(form_headers)]  # Get values from row 3
 
        for header, value in zip(form_headers, form_values):
            form_data[header] = str(value) if value is not None else ''
 
        # Process table fields
        table_start_row = 6  # Table headers start at row 6
        table_headers = [cell.value for cell in ws[table_start_row] if cell.value]
        rows = []
 
        for row in ws.iter_rows(min_row=table_start_row + 1, max_col=len(table_headers)):
            row_data = {}
            has_data = False
           
            for header, cell in zip(table_headers, row):
                value = cell.value
               
                # Handle boolean fields
                if header == 'primaryKey':
                    if isinstance(value, bool):
                        value = value
                    elif isinstance(value, (int, float)):
                        value = bool(value)
                    else:
                        value = str(value).lower().strip() in ['true', '1', 'yes', 'y'] if value else False
                elif value is None:
                    value = ''
                else:
                    value = str(value).strip()
                    if value:
                        has_data = True
               
                row_data[header] = value
 
            if has_data:
                rows.append(row_data)
 
        # Map the data to the required format
        mapped_rows = []
        for row in rows:
            mapped_row = {
                'mapdtlid': '',  # This will be empty for new rows
                'fieldName': row.get('fieldName', ''),
                'dataType': row.get('dataType', ''),
                'primaryKey': row.get('primaryKey', False),
                'pkSeq': row.get('pkSeq', ''),
                'nulls': False,  # Default value, modify if your Excel has this field
                'logic': row.get('logic', ''),
                'validator': 'N',  # Default value
                'keyColumn': row.get('keyColumn', ''),
                'valColumn': row.get('valColumn', ''),
                'mapCombineCode': row.get('mapCombineCode', ''),
                'LogicVerFlag': 'N',  # Default value
                'scdType': row.get('scdType', ''),
                'fieldDesc': row.get('fieldDesc', ''),
                'execSequence': row.get('execSequence', '')
            }
            mapped_rows.append(mapped_row)
 
        # Prepare response data
        response_data = {
            'formData': {
                'reference': str(form_data.get('reference', '')).strip(),
                'description': str(form_data.get('description', '')).strip(),
                'mapperId': '',  # This might need to be generated or extracted
                'targetSchema': str(form_data.get('targetSchema', '')).strip(),
                'tableName': str(form_data.get('tableName', '')).strip(),
                'tableType': str(form_data.get('tableType', '')).strip(),
                'freqCode': str(form_data.get('freqCode', '')).strip(),
                'sourceSystem': str(form_data.get('sourceSystem', '')).strip(),
                'bulkProcessRows': form_data.get('bulkProcessRows', '')
            },
            'rows': mapped_rows
        }
 
        return jsonify(response_data)
    except Exception as e:
        print(f"Error in upload_file: {str(e)}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@mapper_bp.route('/get-by-reference/<reference>', methods=['GET'])
def get_by_reference(reference):
    try:
        conn = create_oracle_connection()
        try:
            # Get reference data
            main_result = get_mapping_ref(conn, reference)
           
            if not main_result:
                return jsonify({
                    'exists': False,
                    'message': 'Reference not found or inactive. You can create a new mapping with this reference.'
                }), 404
           
            # Get mapping details
            details_result = get_mapping_details(conn, reference)
           
            # Format response
            form_data = {
                'reference': main_result['MAPREF'],
                'description': main_result['MAPDESC'] or '',
                'mapperId': str(main_result['MAPID']),
                'targetSchema': main_result['TRGSCHM'] or '',
                'tableName': main_result['TRGTBNM'] or '',
                'tableType': main_result['TRGTBTYP'] or '',
                'freqCode': main_result['FRQCD'] or '',
                'sourceSystem': main_result['SRCSYSTM'] or '',
                'bulkProcessRows': main_result['BLKPRCROWS'],
                'isReferenceDisabled': True
            }
           
            # Transform the details result into rows
            rows = []
            for row in details_result:
                row_data = {
                    'mapdtlid': str(row['MAPDTLID']),
                    'mapref': row['MAPREF'] or '',
                    'fieldName': row['TRGCLNM'] or '',
                    'dataType': row['TRGCLDTYP'] or '',
                    'primaryKey': row['TRGKEYFLG'] == 'Y',
                    'pkSeq': str(row['TRGKEYSEQ']) if row['TRGKEYSEQ'] is not None else '',
                    'fieldDesc': row['TRGCLDESC'] or '',
                    'logic': row['MAPLOGIC'] or '',
                    'keyColumn': row['KEYCLNM'] or '',
                    'valColumn': row['VALCLNM'] or '',
                    'mapCombineCode': row['MAPCMBCD'] or '',
                    'execSequence': str(row['EXCSEQ']) if row['EXCSEQ'] is not None else '',
                    'scdType': str(row['SCDTYP']) if row['SCDTYP'] is not None else '',
                    'LogicVerFlag': row['LGVRFYFLG']
                }
                rows.append(row_data)
           
            # If no detail rows exist, provide empty template rows
            if not rows:
                rows = [{
                    'mapdtlid': '',
                    'mapref': reference,
                    'fieldName': '',
                    'dataType': '',
                    'primaryKey': False,
                    'pkSeq': '',
                    'fieldDesc': '',
                    'logic': '',
                    'keyColumn': '',
                    'valColumn': '',
                    'mapCombineCode': '',
                    'execSequence': '',
                    'scdType': '',
                    'LogicVerFlag': ''
                } for _ in range(6)]
           
            response_data = {
                'exists': True,
                'formData': form_data,
                'rows': rows,
                'message': 'Mapping data retrieved successfully'
            }

            return jsonify(response_data)
        finally:
            conn.close()
           
    except Exception as e:
        print(f"Error in get_by_reference: {str(e)}")
        return jsonify({
            'error': 'An error occurred while retrieving the mapping data',
            'details': str(e)
        }), 500

@mapper_bp.route('/save-to-db', methods=['POST'])
def save_to_db():
    try:
        data = request.json
        form_data = data['formData']
        user_id = form_data['username']
        rows = data['rows']
        modified_rows = data.get('modifiedRows', [])  # Track modified rows
        print(form_data)
        print(rows)
       
        conn = create_oracle_connection()
        try:
            # Oracle connections auto-commit unless explicitly started a transaction
            # No need to explicitly begin a transaction
           
            mapid = create_update_mapping(
                conn,
                form_data['reference'],
                form_data['description'],
                form_data['targetSchema'],
                form_data['tableType'],
                form_data['tableName'],
                form_data['freqCode'],
                form_data['sourceSystem'],
                'Y',  # Default to Y
                datetime.datetime.now(),
                'A' , # Default to Active
                form_data['bulkProcessRows'],
                user_id  # Pass the user_id parameter
            )
           
            processed_rows = []
            for row in rows:
                if not row['fieldName'].strip():
                    continue
               
                row_index = rows.index(row)
                if not (row_index in modified_rows or not row.get('mapdtlid')):
                    continue
                   
                # Set LogicVerFlag to "" if not defined
                logic_ver_flag = row.get('LogicVerFlag', "")  # Use "" if LogicVerFlag is not defined
 
                mapdtlid = create_update_mapping_detail(
                    conn,
                    form_data['reference'],
                    row['fieldName'],
                    row['dataType'],
                    'Y' if row['primaryKey'] else 'N',
                    row['pkSeq'] if row['pkSeq'] and row['primaryKey'] else None,
                    row['fieldDesc'],  # Optional description
                    row['logic'] if row['logic'].strip() else None,
                    row['keyColumn'],
                    row['valColumn'],
                    row['mapCombineCode'],
                    row['execSequence'], 
                    row['scdType'],  
                    logic_ver_flag, 
                    "" if logic_ver_flag == "" else datetime.datetime.now(),  # Set to current time if not empty
                    form_data['username']
                )
               
                processed_rows.append({
                    'index': row_index,
                    'mapdtlid': mapdtlid,
                    'fieldName': row['fieldName']
                })
           
            # Commit the transaction
            conn.commit()
           
            return jsonify({
                'success': True,
                'message': 'Mapping saved successfully',
                'mapperId': str(mapid),
                'processedRows': processed_rows
            })
           
        except Exception as e:
            # Rollback in case of error
            conn.rollback()
            raise e
        finally:
            # Close the connection
            conn.close()
 
    except oracledb.DatabaseError as e:
        # Extract the specific error message that starts with " " and ends with "."
        error_message = str(e)
        match = re.search(r'(?<=::)(.*?)(?=\.)', error_message)  # Look for text between "::" and "."
        specific_error = match.group(0).strip() if match else error_message  # Fallback to the full error message if no match
 
        return jsonify({
            'error': specific_error,
            'details': error_message
        }), 500
   
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"Error in save_to_db: {str(e)}\n{tb}")
        return jsonify({
            'error': f'An error occurred while saving the mapping data {str(e)}',
            'details': str(e)
        }), 500

@mapper_bp.route('/validate-logic', methods=['POST'])
def validate_logic():
    try:
        data = request.json
        p_logic = data.get('p_logic')
        p_keyclnm = data.get('p_keyclnm')
        p_valclnm = data.get('p_valclnm')
       
        if not all([p_logic, p_keyclnm, p_valclnm]):
            return jsonify({
                'status': 'error',
                'message': 'Missing required parameters. Please provide p_logic, p_keyclnm, and p_valclnm.'
            }), 400
       
        connection = None
        try:
            connection = create_oracle_connection()
           
            is_valid, error = validate_logic2(connection, p_logic, p_keyclnm, p_valclnm)
           
            return jsonify({
                'status': 'success',
                'is_valid': is_valid,
                'message': 'Logic is valid' if is_valid == 'Y' else error
            })
        except Exception as e:
            return jsonify({
                'status': 'error',
                'message': f'Error validating logic: {str(e)}'
            }), 500
        finally:
            if connection:
                connection.close()
               
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Error processing request: {str(e)}'
        }), 500

@mapper_bp.route('/validate-batch', methods=['POST'])
def validate_batch_logic():
    try:
        data = request.json
        p_mapref = data.get('mapref')
        rows = data.get('rows', [])
       
        connection = create_oracle_connection()
        try:
            results = []
           
            # First validate all logic together
            bulk_result, bulk_error = validate_all_mapping_details(connection, p_mapref)
            print(bulk_error)
            print(bulk_result)
           
            if bulk_error is None:
                return jsonify({
                    'status': 'success',
                    'bulkValidation': {
                        'success': True,
                        'error': None
                    },
                    'rowResults': results
                })
            

            # Collect all mapdtlids for batch error message retrieval
            map_detail_ids = [row.get('mapdtlid') for row in rows if row.get('mapdtlid') and row.get('logic')]
           
            # Get all error messages in a single batch
            error_messages = get_error_messages_list(connection, map_detail_ids)
           
            # Process results using error messages from get_error_messages_list
            for row in rows:
                if not row.get('logic'):
                    continue
               
                error_message = None
                if row.get('mapdtlid') and row.get('mapdtlid') in error_messages:
                    error_message = error_messages[row.get('mapdtlid')]
                   
                results.append({
                    'rowId': row.get('mapdtlid'),
                    'fieldName': row.get('fieldName'),
                    'isValid': error_message == "Logic is Verified",
                    'error': None if error_message == "Logic is Verified" else error_message,
                    'detailedError': error_message
                })
           
            return jsonify({
                'status': 'success',
                'bulkValidation': {
                    'success': bulk_result,
                    'error': bulk_error
                },
                'rowResults': results
            })
           
        finally:
            connection.close()
           
    except Exception as e:
        traceback.print_exc()
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@mapper_bp.route('/get-parameter-mapping-datatype', methods=['GET'])
def get_parameter_mapping_datatype_api():
    try:
        conn = create_oracle_connection()
        return jsonify(get_parameter_mapping_datatype(conn))
    except Exception as e:
        print(f"Error in get_parameter_mapping_datatype: {str(e)}")
        return jsonify({'error': str(e)}), 500
 
@mapper_bp.route('/parameter_scd_type', methods=["GET"])
def parameter_scd_type():
    try:
        conn = create_oracle_connection()
        try:
            parameter_data = get_parameter_mapping_scd_type(conn)
            return jsonify(parameter_data)
        finally:
            conn.close()
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@mapper_bp.route('/activate-deactivate', methods=['POST'])
def activate_deactivate_mapping():
    try:
        data = request.json
        p_mapref = data.get('mapref')
        p_stflg = data.get('statusFlag')
        
        if not p_mapref or not p_stflg:
            return jsonify({
                'success': False,
                'message': 'Missing required parameters. Please provide mapref and statusFlag.'
            }), 400
            
        if p_stflg not in ['A', 'N']:
            return jsonify({
                'success': False,
                'message': 'Invalid status flag. Must be either "A" (activate) or "N" (deactivate).'
            }), 400
            
        conn = create_oracle_connection()
        try:
            success, message = call_activate_deactivate_mapping(conn, p_mapref, p_stflg)
            return jsonify({
                'success': success,
                'message': message
            })
        finally:
            conn.close()
            
    except Exception as e:
        print(f"Error in activate_deactivate_mapping: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'An error occurred while processing the request: {str(e)}'
        }), 500
    


## By Rajesh

# Get all the mapper reference details
@mapper_bp.route('/get-all-mapper-reference', methods=['GET'])
def get_all_mapper_reference():
    try:
        conn = create_oracle_connection()
        query="""
        SELECT MAPREF, MAPDESC,TRGSCHM,TRGTBTYP,FRQCD,SRCSYSTM,LGVRFYFLG,STFLG,CRTDBY,UPTDBY
        FROM DWMAPR
        WHERE CURFLG = 'Y'
        """
        cursor = conn.cursor()
        cursor.execute(query)
        result = cursor.fetchall()
        cursor.close()
        return jsonify(result)
    except Exception as e:
        print(f"Error in get_all_mapper_reference: {str(e)}")
        return jsonify({'error': str(e)}), 500




# Delete the mapper reference
@mapper_bp.route('/delete-mapper-reference', methods=['POST'])
def delete_mapper_reference():
    try:
        data = request.json
        p_mapref = data.get('mapref')
        conn = create_oracle_connection()
        
        try:
            # Call the Oracle procedure instead of executing the query directly
            success, message = call_delete_mapping(conn, p_mapref)
            
            if success:
                return jsonify({
                    'success': True,
                    'message': message
                })
            else:
                return jsonify({
                    'success': False,
                    'message': message
                }), 400
        finally:
            conn.close()
    except Exception as e:
        print(f"Error in delete_mapper_reference: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Delete mapping detail row
@mapper_bp.route('/delete-mapping-detail', methods=['POST'])
def delete_mapping_detail():
    try:
        data = request.json
        p_mapref = data.get('mapref')
        p_trgclnm = data.get('trgclnm')
        
        if not p_mapref or not p_trgclnm:
            return jsonify({
                'success': False,
                'message': 'Missing required parameters. Please provide mapref and trgclnm.'
            }), 400
            
        conn = create_oracle_connection()
        try:
            # Call the Oracle procedure through our helper function
            success, message = call_delete_mapping_details(conn, p_mapref, p_trgclnm)
            
            return jsonify({
                'success': success,
                'message': message
            })
        finally:
            conn.close()
            
    except Exception as e:
        print(f"Error in delete_mapping_detail: {str(e)}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'An error occurred while deleting the mapping detail: {str(e)}'
        }), 500


# Schedule the job immediately
@mapper_bp.route('/schedule-job-immediately', methods=['POST'])
def schedule_job_immediately():
    try:
        data = request.json
        p_mapref = data.get('mapref')

        if not p_mapref:
            return jsonify({
                'success': False,
                'message': 'Missing required parameter: mapref'
            }), 400

        conn = create_oracle_connection()
        try:
            success, message = call_schedule_immediate_job(conn, p_mapref)
            return jsonify({
                'success': success,
                'message': message  
            })
        finally:
            conn.close()
    except Exception as e:
        print(f"Error in schedule_job_immediately: {str(e)}")
        return jsonify({'error': str(e)}), 500
    
